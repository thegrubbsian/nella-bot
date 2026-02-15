"""Text extraction from binary document files (PDF, DOCX, XLSX)."""

import asyncio
import logging
import mimetypes
from pathlib import Path

logger = logging.getLogger(__name__)

MAX_EXTRACTED_CHARS = 50_000


def _extract_pdf(file_path: Path) -> str | None:
    """Extract text from a PDF file using PyMuPDF."""
    import pymupdf

    doc = pymupdf.open(file_path)
    try:
        pages = []
        for page in doc:
            text = page.get_text()
            if text.strip():
                pages.append(text.strip())
        return "\n\n".join(pages) if pages else None
    finally:
        doc.close()


def _extract_docx(file_path: Path) -> str | None:
    """Extract text from a DOCX file using python-docx."""
    import docx

    doc = docx.Document(file_path)
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs) if paragraphs else None


def _extract_xlsx(file_path: Path) -> str | None:
    """Extract text from an XLSX file using openpyxl, rendered as CSV."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheets = wb.sheetnames
        multi_sheet = len(sheets) > 1
        parts: list[str] = []

        for name in sheets:
            ws = wb[name]
            lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(",".join(cells))
            if lines:
                if multi_sheet:
                    parts.append(f"## Sheet: {name}\n" + "\n".join(lines))
                else:
                    parts.append("\n".join(lines))

        return "\n\n".join(parts) if parts else None
    finally:
        wb.close()


_EXTRACTORS: dict[str, callable] = {
    "application/pdf": _extract_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": _extract_docx,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": _extract_xlsx,
}


async def extract_text(file_path: Path) -> str | None:
    """Extract text from a document file.

    Returns extracted text (truncated to MAX_EXTRACTED_CHARS), or None if the
    format is unsupported or extraction fails.
    """
    mime_type = mimetypes.guess_type(file_path)[0]
    if not mime_type:
        return None

    extractor = _EXTRACTORS.get(mime_type)
    if not extractor:
        return None

    try:
        text = await asyncio.to_thread(extractor, file_path)
    except Exception:
        logger.exception("Failed to extract text from %s", file_path.name)
        return None

    if not text:
        return None

    if len(text) > MAX_EXTRACTED_CHARS:
        text = text[:MAX_EXTRACTED_CHARS] + "\n\n[truncated — extracted text too large]"

    return text
